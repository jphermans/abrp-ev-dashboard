"""ABRP Excel export parser — converts .xlsx to dashboard records."""

from datetime import datetime
from pathlib import Path

PROVIDERS = [
    ("DATS 24", ["dats 24", "dats24"]),
    ("Fastned", ["fastned"]),
    ("Allego", ["allego"]),
    ("Shell Recharge", ["shell recharge", "shell"]),
    ("Ionity", ["ionity"]),
    ("PluginCompany", ["plugincompany", "plugin company"]),
    ("T-Line", ["t-line"]),
    ("Electra", ["electra"]),
    ("EVBox", ["evbox"]),
    ("Lidl", ["lidl"]),
]

WEEKDAYS_NL = ["Maandag", "Dinsdag", "Woensdag", "Donderdag", "Vrijdag", "Zaterdag", "Zondag"]
WEEKDAYS_EN = ["Monday", "Tuesday", "Wednesday", "Thursday", "Friday", "Saturday", "Sunday"]


def _extract_provider(text):
    if not text:
        return None
    low = text.lower()
    for name, patterns in PROVIDERS:
        for p in patterns:
            if p in low:
                return name
    return None


def _extract_location(loc_text):
    if not loc_text:
        return None
    parts = str(loc_text).strip().split("\n", 1)
    if len(parts) > 1:
        return parts[1].strip().strip("()")
    return None


def _safe_float(val):
    """Safely convert a cell value to float, returning None on failure."""
    if val is None:
        return None
    try:
        return float(val)
    except (ValueError, TypeError):
        return None


def _safe_int_pct(val):
    """Safely convert a 0-1 fraction to an integer percentage."""
    f = _safe_float(val)
    if f is None:
        return None
    return round(f * 100)


def parse_excel_to_records(filepath):
    """Parse ABRP Excel export to dashboard records."""
    import openpyxl

    wb = openpyxl.load_workbook(filepath, data_only=True, read_only=True)
    ws = wb.active
    records = []

    for row in ws.iter_rows(min_row=4, max_row=ws.max_row, values_only=True):
        # Guard against short/malformed rows (M2)
        if not row or row[0] is None:
            continue
        if len(row) < 13:
            continue

        activity = row[0]
        start_raw = row[1]

        if isinstance(start_raw, datetime):
            dt = start_raw
        elif isinstance(start_raw, str):
            try:
                dt = datetime.strptime(start_raw, "%m/%d/%Y %H:%M")
            except ValueError:
                continue
        else:
            continue

        distance_mi = _safe_float(row[4])
        distance_km = round(distance_mi * 1.609344, 1) if distance_mi is not None else None
        energy = _safe_float(row[9])

        # Provider detection: use END location (row[6]) where the charger actually is,
        # NOT the combined start+end text (which causes false matches when the car
        # was at a provider location before driving elsewhere to charge)
        start_loc = str(row[5] or '')
        end_loc = str(row[6] or '')

        if activity == "Laad op":
            # Provider: detect from END location ONLY (where the charger is).
            # Do NOT fall back to START — that's the previous destination,
            # which may contain a different provider name and cause false attribution.
            charge_provider = _extract_provider(end_loc)
            charge_location = _extract_location(end_loc) or _extract_location(start_loc)
        else:
            charge_provider = None
            charge_location = None

        records.append({
            "date": dt.strftime("%Y-%m-%d"),
            "time": dt.strftime("%H:%M"),
            "datetime": dt.strftime("%Y-%m-%dT%H:%M"),
            "weekday": WEEKDAYS_NL[dt.weekday()],
            "activity": activity,
            "duration": str(row[3] or ""),
            "distance_km": distance_km,
            "distance_mi": distance_mi,
            "start_soc": _safe_int_pct(row[7]),
            "end_soc": _safe_int_pct(row[8]),
            "energy_kwh": energy,
            "start_odo_mi": _safe_float(row[10]),
            "end_odo_mi": _safe_float(row[11]),
            "vehicle": str(row[12] or "EV"),
            "charge_provider": charge_provider,
            "charge_location": charge_location,
        })
    wb.close()
    return records


def get_data_file(data_dir: Path):
    """Return path to merged JSON or most recent Excel."""
    json_file = data_dir / "activities.json"
    if json_file.exists():
        return json_file
    excels = sorted(data_dir.glob("*.xlsx"))
    if excels:
        return excels[-1]
    return None
