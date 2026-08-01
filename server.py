#!/usr/bin/env python3
"""
ABRP EV Dashboard Server
Serves the dashboard and proxies ABRP API calls with rate limiting.
Designed for Raspberry Pi 4/5.
"""

import json
import time
import threading
import os
import glob
from datetime import datetime, timedelta
from pathlib import Path

from flask import Flask, send_file, request, jsonify, render_template_string

app = Flask(__name__)

# ─── Config ───────────────────────────────────────────────────────
BASE_DIR = Path(__file__).parent
DATA_DIR = BASE_DIR / "data"
DATA_DIR.mkdir(exist_ok=True)
DASHBOARD_HTML = BASE_DIR / "templates" / "dashboard.html"

# Rate limiter: max 2 requests/second to ABRP API
API_RATE_LIMIT = 0.5  # 500ms between calls (2/sec max)
_last_api_call = [0.0]
_api_lock = threading.Lock()

# Simple in-memory cache
_cache = {}
CACHE_TTL = 300  # 5 minutes


def rate_limited():
    """Enforce 2 requests/second rate limit."""
    with _api_lock:
        elapsed = time.time() - _last_api_call[0]
        if elapsed < API_RATE_LIMIT:
            time.sleep(API_RATE_LIMIT - elapsed)
        _last_api_call[0] = time.time()


# ─── Data Storage ─────────────────────────────────────────────────
def get_data_file():
    """Get the path to the current data file."""
    # Prefer the merged JSON cache (has all records from all files)
    json_file = DATA_DIR / "activities.json"
    if json_file.exists():
        return json_file
    # No JSON yet? Check for Excel files
    excel_files = sorted(DATA_DIR.glob("*.xlsx"))
    if excel_files:
        return excel_files[-1]  # Most recent
    return None


def parse_excel_to_records(filepath):
    """Parse ABRP Excel export to dashboard records."""
    try:
        import openpyxl
    except ImportError:
        os.system(f"pip3 install openpyxl")
        import openpyxl

    wb = openpyxl.load_workbook(filepath, data_only=True)
    ws = wb.active

    PROVIDERS = [
        ("DATS 24", ["dats 24", "dats24"]),
        ("Fastned", ["fastned"]),
        ("Allego", ["allego"]),
        ("Shell Recharge", ["shell recharge", "shell"]),
        ("Ionity", ["ionity"]),
        ("PluginCompany", ["plugincompany", "plugin company"]),
        ("T-Line", ["t-line"]),
        ("Electra", ["electra"]),
        ("EVBox", ["evbox"]),
        ("Lidl", ["lidl"]),
    ]

    def extract_provider(text):
        if not text:
            return None
        low = text.lower()
        for name, patterns in PROVIDERS:
            for p in patterns:
                if p in low:
                    return name
        return None

    def extract_location(loc_text):
        if not loc_text:
            return None
        parts = str(loc_text).strip().split("\n", 1)
        if len(parts) > 1:
            return parts[1].strip().strip("()")
        return None

    records = []
    for row in ws.iter_rows(min_row=4, max_row=ws.max_row, values_only=True):
        if not row[0]:
            continue
        activity = row[0]
        start_raw = row[1]

        # Parse datetime
        if isinstance(start_raw, datetime):
            dt = start_raw
        elif isinstance(start_raw, str):
            try:
                dt = datetime.strptime(start_raw, "%m/%d/%Y %H:%M")
            except ValueError:
                continue
        else:
            continue

        distance_mi = row[4]
        distance_km = round(float(distance_mi) * 1.609344, 1) if distance_mi else None
        energy = float(row[9]) if row[9] else None
        all_loc = f"{row[5] or ''} {row[6] or ''}"

        record = {
            "date": dt.strftime("%Y-%m-%d"),
            "time": dt.strftime("%H:%M"),
            "datetime": dt.strftime("%Y-%m-%dT%H:%M"),
            "weekday": ["Maandag", "Dinsdag", "Woensdag", "Donderdag", "Vrijdag", "Zaterdag", "Zondag"][dt.weekday()],
            "activity": activity,
            "duration": row[3] or "",
            "distance_km": distance_km,
            "distance_mi": float(distance_mi) if distance_mi else None,
            "start_soc": round(float(row[7]) * 100) if row[7] is not None else None,
            "end_soc": round(float(row[8]) * 100) if row[8] is not None else None,
            "energy_kwh": energy,
            "start_odo_mi": float(row[10]) if row[10] else None,
            "end_odo_mi": float(row[11]) if row[11] else None,
            "vehicle": row[12] or "EV",
            "charge_provider": extract_provider(all_loc) if activity == "Laad op" else None,
            "charge_location": extract_location(row[5] or row[6]) if activity == "Laad op" else None,
        }
        records.append(record)

    return records


# ─── Routes ───────────────────────────────────────────────────────
@app.route("/")
def index():
    """Serve the dashboard."""
    return send_file(DASHBOARD_HTML)


@app.route("/api/data")
def get_data():
    """Get all stored activity data as JSON."""
    cache_key = "all_data"
    cached = _cache.get(cache_key)
    if cached and time.time() - cached["time"] < CACHE_TTL:
        return jsonify(cached["data"])

    data_file = get_data_file()
    if not data_file:
        return jsonify([])

    if str(data_file).endswith(".json"):
        with open(data_file) as f:
            data = json.load(f)
    elif str(data_file).endswith(".xlsx"):
        data = parse_excel_to_records(data_file)
        # Cache as JSON
        with open(DATA_DIR / "activities.json", "w") as f:
            json.dump(data, f, ensure_ascii=False)
    else:
        data = []

    _cache[cache_key] = {"time": time.time(), "data": data}
    return jsonify(data)


@app.route("/api/upload", methods=["POST"])
def upload_excel():
    """Upload one or more ABRP Excel exports."""
    if "files" not in request.files:
        return jsonify({"error": "No files provided"}), 400

    files = request.files.getlist("files")
    all_records = []

    # Load existing data
    existing_json = DATA_DIR / "activities.json"
    if existing_json.exists():
        with open(existing_json) as f:
            all_records = json.load(f)

    uploaded = 0
    for file in files:
        if not file.filename or not file.filename.endswith((".xlsx", ".xls")):
            continue
        filepath = DATA_DIR / file.filename
        file.save(filepath)
        records = parse_excel_to_records(filepath)
        all_records.extend(records)
        uploaded += 1

    # Deduplicate
    seen = set()
    deduped = []
    for r in all_records:
        key = f"{r['datetime']}|{r['activity']}"
        if key not in seen:
            seen.add(key)
            deduped.append(r)
    deduped.sort(key=lambda x: x["datetime"])

    # Save merged data
    with open(existing_json, "w") as f:
        json.dump(deduped, f, ensure_ascii=False)

    # Clear cache
    _cache.clear()

    return jsonify({
        "status": "ok",
        "uploaded": uploaded,
        "total_records": len(deduped),
        "message": f"{uploaded} file(s) processed, {len(deduped)} total records"
    })


@app.route("/api/abrp/login", methods=["POST"])
def abrp_login():
    """
    Login to ABRP and fetch activities.
    Requires ABRP account credentials and API key with 'session' feature.
    """
    data = request.json or {}
    email = data.get("email", "")
    password = data.get("password", "")
    api_key = data.get("api_key", "")

    if not email or not password or not api_key:
        return jsonify({"error": "email, password, and api_key are required"}), 400

    try:
        import urllib.request
        import urllib.error

        rate_limited()

        # Step 1: Login to get session token
        login_url = "https://api.iternio.com/2/auth/login"
        login_body = json.dumps({"username": email, "password": password}).encode()
        req = urllib.request.Request(login_url, data=login_body, method="POST")
        req.add_header("Content-Type", "application/json")
        req.add_header("X-API-KEY", api_key)

        try:
            with urllib.request.urlopen(req, timeout=15) as resp:
                login_resp = json.loads(resp.read())
        except urllib.error.HTTPError as e:
            error_body = e.read().decode()
            return jsonify({
                "error": f"ABRP login failed (HTTP {e.code})",
                "details": error_body[:500],
                "hint": "Your API key may need the 'session' feature (premium plan required)"
            }), 502

        session_token = login_resp.get("access_token")
        if not session_token:
            return jsonify({"error": "No access_token in login response"}), 502

        # Step 2: Get vehicle list
        rate_limited()
        vehicles_url = "https://api.iternio.com/2/vehicle/_list"
        veh_req = urllib.request.Request(vehicles_url, data=b"{}", method="POST")
        veh_req.add_header("Content-Type", "application/json")
        veh_req.add_header("X-API-KEY", api_key)
        veh_req.add_header("X-ABRP-SESSION", session_token)

        try:
            with urllib.request.urlopen(veh_req, timeout=15) as resp:
                vehicles = json.loads(resp.read())
        except urllib.error.HTTPError as e:
            return jsonify({
                "error": f"Failed to get vehicle list (HTTP {e.code})",
                "details": e.read().decode()[:500]
            }), 502

        return jsonify({
            "status": "ok",
            "session_token": session_token,
            "vehicles": vehicles,
            "message": "Login successful. Use /api/abrp/activities to fetch data."
        })

    except Exception as e:
        return jsonify({"error": str(e)}), 500


@app.route("/api/abrp/activities", methods=["POST"])
def abrp_activities():
    """Fetch activities from ABRP API (requires session token)."""
    data = request.json or {}
    api_key = data.get("api_key", "")
    session_token = data.get("session_token", "")
    vehicle_id = data.get("vehicle_id")
    start_time = data.get("start_time")  # unix timestamp
    end_time = data.get("end_time")      # unix timestamp

    if not all([api_key, session_token, vehicle_id, start_time, end_time]):
        return jsonify({"error": "api_key, session_token, vehicle_id, start_time, end_time required"}), 400

    try:
        import urllib.request
        import urllib.error

        rate_limited()

        url = "https://api.iternio.com/2/session/get_tlm_activities"
        body = json.dumps({
            "vehicle_id": vehicle_id,
            "start_time": int(start_time),
            "end_time": int(end_time)
        }).encode()

        req = urllib.request.Request(url, data=body, method="POST")
        req.add_header("Content-Type", "application/json")
        req.add_header("X-API-KEY", api_key)
        req.add_header("X-ABRP-SESSION", session_token)

        with urllib.request.urlopen(req, timeout=30) as resp:
            activities = json.loads(resp.read())

        return jsonify({"status": "ok", "activities": activities})

    except urllib.error.HTTPError as e:
        return jsonify({
            "error": f"ABRP API error (HTTP {e.code})",
            "details": e.read().decode()[:500]
        }), 502
    except Exception as e:
        return jsonify({"error": str(e)}), 500


@app.route("/api/vw/sync", methods=["POST"])
def vw_sync():
    """
    Fetch trip and charging data from Volkswagen WeConnect.
    Uses the weconnect Python library (free, no API key needed).
    Requires VW account credentials (email + password).
    """
    data = request.json or {}
    username = data.get("username", "")
    password = data.get("password", "")
    spin = data.get("spin", "")

    # Also check environment / config file
    if not username:
        username = os.environ.get("VW_USERNAME", "")
    if not password:
        password = os.environ.get("VW_PASSWORD", "")

    if not username or not password:
        return jsonify({"error": "VW username (email) and password required"}), 400

    try:
        from weconnect import WeConnect
        from weconnect.domain import Domain
    except ImportError:
        return jsonify({
            "error": "weconnect library not installed",
            "hint": "Install with: pip install weconnect"
        }), 500

    try:
        rate_limited()

        # Token cache file (avoids re-login on every sync)
        token_file = str(DATA_DIR / "vw_token.json")

        wc = WeConnect(
            username=username,
            password=password,
            spin=spin or None,
            tokenfile=token_file,
            updateAfterLogin=True,
            loginOnInit=True,
            updatePictures=False,
            maxAge=300,  # 5 min cache
            timeout=30,
            numRetries=2,
            selective=[Domain.TRIPS, Domain.CHARGING, Domain.MEASUREMENTS, Domain.BATTERY_SUPPORT],
        )

        records = []

        # Iterate vehicles
        for vin, vehicle in wc.vehicles.items():
            vehicle_name = getattr(vehicle, 'nickname', None) or vin

            # ─── Trips ───
            try:
                trips_dict = vehicle.getByAddressString("trips", allowEmpty=True) if vehicle.statusExists("trips") else None
                if trips_dict:
                    for trip_id, trip in trips_dict.items():
                        trip_data = trip.asDict() if hasattr(trip, 'asDict') else {}
                        distance_km = trip_data.get('mileage_km', 0) - trip_data.get('startMileage_km', 0)
                        if distance_km < 0:
                            distance_km = trip_data.get('mileage_km', 0)

                        records.append({
                            "date": str(trip_data.get('tripEndTimestamp', ''))[:10],
                            "time": str(trip_data.get('tripEndTimestamp', ''))[11:16],
                            "datetime": str(trip_data.get('tripEndTimestamp', '')).replace(' ', 'T')[:16],
                            "weekday": "",
                            "activity": "Rijd",
                            "duration": f"{trip_data.get('travelTime', 0)} sec",
                            "distance_km": round(distance_km, 1) if distance_km else None,
                            "distance_mi": round(distance_km / 1.609344, 1) if distance_km else None,
                            "start_soc": None,
                            "end_soc": None,
                            "energy_kwh": None,
                            "start_odo_mi": round(trip_data.get('startMileage_km', 0) / 1.609344, 1) if trip_data.get('startMileage_km') else None,
                            "end_odo_mi": round(trip_data.get('mileage_km', 0) / 1.609344, 1) if trip_data.get('mileage_km') else None,
                            "vehicle": str(vehicle_name),
                            "charge_provider": None,
                            "charge_location": None,
                            "_avg_speed": trip_data.get('averageSpeed_kmph'),
                            "_avg_consumption": trip_data.get('averageElectricConsumption'),
                            "_avg_recuperation": trip_data.get('averageRecuperation'),
                        })
            except Exception as e:
                print(f"  ⚠️ Trips error: {e}")

            # ─── Charging sessions ───
            try:
                if vehicle.statusExists("charging"):
                    charging = vehicle.getByAddressString("charging", allowEmpty=True)
                    if charging:
                        charge_data = charging.asDict() if hasattr(charging, 'asDict') else {}
                        # Current charging status (not historical sessions from WeConnect basic API)
                        # Historical charging sessions come from the trips endpoint
                        records.append({
                            "date": "",
                            "time": "",
                            "datetime": "",
                            "weekday": "",
                            "activity": "Laad op",
                            "duration": "",
                            "distance_km": None,
                            "distance_mi": None,
                            "start_soc": round(charge_data.get('batteryStatus', {}).get('currentSOC_pct', 0)) if charge_data.get('batteryStatus') else None,
                            "end_soc": None,
                            "energy_kwh": None,
                            "vehicle": str(vehicle_name),
                            "charge_provider": None,
                            "charge_location": charge_data.get('chargingStatus', {}).get('chargeType', '') if charge_data.get('chargingStatus') else None,
                        })
            except Exception as e:
                print(f"  ⚠️ Charging error: {e}")

        wc.disconnect()

        if not records:
            return jsonify({
                "status": "ok",
                "message": "Login successful but no trip data found. Your vehicle may not expose trip history via WeConnect, or terms need to be accepted in the VW app first.",
                "records": 0
            })

        # Merge with existing data
        existing_json = DATA_DIR / "activities.json"
        all_records = []
        if existing_json.exists():
            with open(existing_json) as f:
                all_records = json.load(f)

        all_records.extend(records)

        # Deduplicate
        seen = set()
        deduped = []
        for r in all_records:
            key = f"{r.get('datetime','')}|{r.get('activity','')}"
            if key and key not in seen:
                seen.add(key)
                deduped.append(r)
        deduped.sort(key=lambda x: x.get("datetime", ""))

        with open(existing_json, "w") as f:
            json.dump(deduped, f, ensure_ascii=False)

        _cache.clear()

        return jsonify({
            "status": "ok",
            "message": f"WeConnect sync complete: {len(records)} records fetched, {len(deduped)} total",
            "fetched": len(records),
            "total": len(deduped)
        })

    except Exception as e:
        error_msg = str(e)
        if "401" in error_msg or "Unauthorized" in error_msg:
            return jsonify({"error": "VW login failed — check your email and password"}), 401
        if "terms" in error_msg.lower():
            return jsonify({"error": "Please accept WeConnect terms in the VW app first"}), 403
        return jsonify({"error": error_msg[:500]}), 500


@app.route("/api/vw/status", methods=["GET"])
def vw_status():
    """Check if WeConnect is configured."""
    config_file = DATA_DIR / "vw_config.json"
    if config_file.exists():
        with open(config_file) as f:
            config = json.load(f)
        return jsonify({
            "configured": True,
            "username": config.get("username", ""),
            "has_password": bool(config.get("password")),
            "last_sync": config.get("last_sync"),
        })
    return jsonify({"configured": bool(os.environ.get("VW_USERNAME"))})


@app.route("/api/vw/config", methods=["POST"])
def vw_config():
    """Save or delete WeConnect credentials."""
    data = request.json or {}
    action = data.get("action", "save")
    config_file = DATA_DIR / "vw_config.json"

    if action == "delete":
        config_file.unlink(missing_ok=True)
        (DATA_DIR / "vw_token.json").unlink(missing_ok=True)
        return jsonify({"status": "ok", "message": "WeConnect credentials deleted"})

    username = data.get("username", "").strip()
    password = data.get("password", "").strip()
    spin = data.get("spin", "").strip()

    if not username or not password:
        return jsonify({"error": "Username and password required"}), 400

    config = {
        "username": username,
        "password": password,
        "spin": spin,
        "last_sync": None,
    }
    with open(config_file, "w") as f:
        json.dump(config, f, ensure_ascii=False)

    return jsonify({"status": "ok", "message": "WeConnect credentials saved"})


@app.route("/api/status")
def status():
    """Health check endpoint."""
    data_file = get_data_file()
    excel_files = list(DATA_DIR.glob("*.xlsx"))
    json_file = DATA_DIR / "activities.json"

    record_count = 0
    if json_file.exists():
        try:
            with open(json_file) as f:
                record_count = len(json.load(f))
        except:
            pass

    # Check WeConnect config
    vw_configured = bool(os.environ.get("VW_USERNAME") or (DATA_DIR / "vw_config.json").exists())

    return jsonify({
        "status": "running",
        "data_source": str(data_file) if data_file else "none",
        "excel_files": len(excel_files),
        "total_records": record_count,
        "cache_entries": len(_cache),
        "pi_model": get_pi_model(),
        "vw_connected": vw_configured,
    })


def get_pi_model():
    """Detect Raspberry Pi model."""
    try:
        with open("/proc/device-tree/model") as f:
            return f.read().strip().strip("\x00")
    except:
        return "unknown"


# ─── Main ─────────────────────────────────────────────────────────
if __name__ == "__main__":
    port = int(os.environ.get("PORT", 8000))
    host = os.environ.get("HOST", "0.0.0.0")
    print(f"🚗 ABRP Dashboard starting on http://{host}:{port}")
    print(f"   Pi model: {get_pi_model()}")
    print(f"   Data dir: {DATA_DIR}")

    # Auto-import any Excel files in data/ on startup
    excel_files = list(DATA_DIR.glob("*.xlsx"))
    if excel_files:
        print(f"   Found {len(excel_files)} Excel file(s), importing...")
        all_records = []
        for ef in excel_files:
            recs = parse_excel_to_records(ef)
            all_records.extend(recs)
            print(f"   → {ef.name}: {len(recs)} records")
        # Deduplicate
        seen = set()
        deduped = []
        for r in all_records:
            k = f"{r['datetime']}|{r['activity']}"
            if k not in seen:
                seen.add(k)
                deduped.append(r)
        deduped.sort(key=lambda x: x["datetime"])
        with open(DATA_DIR / "activities.json", "w") as f:
            json.dump(deduped, f, ensure_ascii=False)
        print(f"   ✅ {len(deduped)} records loaded")

    app.run(host=host, port=port, debug=False)
